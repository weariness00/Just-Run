from openpyxl import load_workbook
from Scripts.Object.Monster.MonsterPool import MonsterPool

class LevelingData:
    def __init__(self, level):
        self.file = load_workbook('Data/LevelingData.xlsx', data_only = True)
        self.level = self.file[level]
        self.maxLevel = self.level["K2"].value
        pass

    def __del__(self):
        pass

    def SetTimeData(self, playTimer):
        playTimer.winTime = self.level["L2"].value
        playTimer.levelUpLengthTime = self.level["M2"].value
        pass
    def SetDefaultPoolData(self, pools):
        maxPool = self.level["N2"].value
        for i in range(maxPool):
            type = self.level.cell(row=i + 2, column=1).value
            maxCount = self.level.cell(row=i + 2, column=2).value
            sapwnCount = self.level.cell(row=i + 2, column=3).value
            coolTime = self.level.cell(row=i + 2, column=4).value
            pools.append(MonsterPool(type, maxCount, sapwnCount, coolTime))
            pass
        pass

    def SetPoolData(self, pools, level):
        if level > self.maxLevel:
            level = self.maxLevel

        r = 22
        c = level * 3
        for pool in pools:
            pool.spawnCount += self.level.cell(row=r, column=c).value
            pool.coolTime += self.level.cell(row=r, column=c + 1).value
            r += 1
            pass
        pass