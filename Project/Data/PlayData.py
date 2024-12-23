from openpyxl import load_workbook
from Scripts.Object.Monster.Monster import Monster
from Scripts.Object.Item.Item import Item

class PlayData:
    def __init__(self):
        self.file = load_workbook('Data/PlayData.xlsx', data_only = True)
        self.sheet = self.file["Sheet1"]
        pass

    def __del__(self):
        pass

    def SetData(self, game):
        pass

    def GetData(self, game):
        nextIndex = self.sheet["A1"].value
        for i in range(3, 8):
            self.sheet.cell(row = nextIndex + 2, column=i).value = 0
        self.sheet.cell(row=nextIndex + 2, column=2).value = "Unknown"
        self.sheet.cell(row=nextIndex + 2, column=3).value = game.playTimer.nowTime
        self.sheet.cell(row=nextIndex + 2, column=4).value = Monster.deathCount
        self.sheet.cell(row=nextIndex + 2, column=5).value = Item.earnCount

        # Player 정보 저장
        self.sheet.cell(row=nextIndex + 2, column=8).value = game.player.name
        self.sheet.cell(row=nextIndex + 2, column=9).value = game.player.speed
        self.sheet.cell(row=nextIndex + 2, column=10).value = game.player.maxLife
        self.sheet.cell(row=nextIndex + 2, column=11).value = game.player.life

        score = Monster.deathCount * 3 + Item.earnCount * 5
        if game.player.name.find("Easy") != -1:
            score += game.playTimer.nowTime + \
                     game.player.speed + \
                     game.player.maxLife + \
                     game.player.life * 10
        elif game.player.name.find("Normal") != -1:
            score += game.playTimer.nowTime * 5 + \
                     game.player.speed * 1.3 + \
                     game.player.maxLife * 1.3 + \
                     game.player.life * 2 * 10
        elif game.player.name.find("Hard") != -1:
            score += game.playTimer.nowTime * 10 + \
                     game.player.speed * 2 + \
                     game.player.maxLife * 2 + \
                     game.player.life * 3 * 10
        self.sheet.cell(row=nextIndex + 2, column=12).value = score

        self.sheet["A1"].value = nextIndex + 1
        self.file.save('Data/PlayData.xlsx')
        pass

    def GetPlayData(self, score):
        lastScoreNum = self.sheet["A1"].value
        t = self.sheet.cell(row=lastScoreNum + 1, column=3).value.__str__()
        if len(t) > 2:
            ts = t.split()
            ts.insert(3, ' : ')
            t = ' '.join(ts)
        timeText = "[PlayTime]" + t
        score.playTimeText.SetTextBox(10, 2, timeText)
        score.killCountText.SetTextBox(12,2, "[Kill Count]" + self.sheet.cell(row=lastScoreNum + 1, column=4).value.__str__())
        score.earnItemCountText.SetTextBox(17,2, "[Earn Item Count]" + self.sheet.cell(row=lastScoreNum + 1, column=5).value.__str__())
        score.scoreText.SetTextBox(7, 2, "[Score]" + self.sheet.cell(row=lastScoreNum + 1, column=12).value.__str__())
        pass
