from Scripts.Afx import *
from Scripts.FrameWork.UI import UI
from Scripts.FrameWork.Object import Object
from Scripts.FrameWork.Render import Render
from Scripts.UI.Button import Button
from Scripts.UI.Text import Text
from Scripts.UI.InputFiled import InputFiled

import Scripts.FrameWork.game_framework as game_framework
import Scripts.State.GamePlay_State as game_state
import Scripts.State.RankGraph_State as rank_state

from openpyxl import load_workbook

# Render
ui_Render = None
text_Render = None

updateList = None

def enter():
    global ui_Render, text_Render
    global updateList

    updateList = []
    Object.updateList = updateList

    ui_Render = Render()
    ui_Render.name += " : UI - Lobby"
    UI.renderList = ui_Render

    text_Render = Render()
    text_Render.name += " : Text - Lobby"
    Text.renderList = text_Render

    #객체 초기화
    background = Object()
    background.image = load_image('image/Background/Forest.png')
    background.image_type = [0, 450, 1024, 450]
    background.transform.Position += Instance.windowSize / 2
    background.transform.Scale *= Instance.windowSize / [background.image_type[2], background.image_type[3]]
    ui_Render.AddObject(background, 0)

    BannerPopUp()
    ModePopUp().SetActive(False)
    pass

# finalization code
def exit():
    global ui_Render, text_Render
    Object.AllObject.clear()
    Object.updateList.clear()

    ui_Render.__del__()
    text_Render.__del__()
    pass

def handle_events():
    events = get_events()
    Object.events = events

    for obj in updateList:
        if not obj.isActive:
            continue
        obj.Handle_Event()

    for event in events:
        if event.type == SDL_QUIT:
            game_framework.quit()

        pass
    pass


def update():
    for obj in updateList:
        if not obj.isActive:
            continue
        obj.Update()

    pass

def event_update():
    for obj in Object.updateList:
        obj.EventCall()
    pass

def draw():

    ui_Render.UIDraw()

    text_Render.TextDraw()

    pass

def pause():
    pass

def resume():
    pass

class BannerPopUp(Object):
    Instance = None
    def __init__(self):
        super(BannerPopUp, self).__init__()
        self.frame = Object()
        self.frame.image = load_image('image/UI/Frame_White.png')
        self.frame.image_type = [0, 0, 600, 600]
        self.frame.transform.Position += Instance.windowSize / 2
        self.frame.transform.Scale *= 1.2

        self.banner = Text(120)
        self.banner.text = '그냥 튀어!'
        self.banner.color = [62, 62, 62]
        self.banner.font = self.banner.fontList['Explain']
        self.banner.transform.Position = Instance.windowSize // 2 + [-230, 200]
        self.subBanner = Text(40)
        self.subBanner.text = 'Just Run!'
        self.subBanner.color = [62, 62, 62]
        self.subBanner.transform.Position = self.banner.transform.Position + [130, -100]

        # Start
        self.start_Button = Button(1)
        self.start_Button.transform.Scale *= 2
        self.start_Button.transform.Position += [Instance.windowSize[0] // 2, 400]
        self.start_Button.text = Text(40)
        self.start_Button.text.text = '시작'
        self.start_Button.text.transform.Position += self.start_Button.transform.Position + [-40, 0]

        # Rank Graph
        self.rank_Graph_Button = Button(1)
        self.rank_Graph_Button.transform.Scale *= 2
        self.rank_Graph_Button.transform.Position += [Instance.windowSize[0] // 2, 300]
        self.rank_Graph_Button.text = Text(40)
        self.rank_Graph_Button.text.text = '랭크'
        self.rank_Graph_Button.text.transform.Position += self.rank_Graph_Button.transform.Position + [-40, 0]

        # Exit
        self.exit_Button = Button(1)
        self.exit_Button.transform.Scale *= 2
        self.exit_Button.transform.Position += [Instance.windowSize[0] // 2, 200]
        self.exit_Button.text = Text(40)
        self.exit_Button.text.text = '종료'
        self.exit_Button.text.transform.Position += self.exit_Button.transform.Position + [-40, 0]
        # Render 초기화
        ui_Render.AddObject(self.frame, 0)

        BannerPopUp.Instance = self
        pass

    def Enable(self):
        self.frame.SetActive(True)
        self.banner.SetActive(True)
        self.subBanner.SetActive(True)
        self.start_Button.SetActive(True)
        self.start_Button.text.SetActive(True)
        self.rank_Graph_Button.SetActive(True)
        self.rank_Graph_Button.text.SetActive(True)
        self.exit_Button.SetActive(True)
        self.exit_Button.text.SetActive(True)
        pass

    def Disable(self):
        self.frame.SetActive(False)
        self.banner.SetActive(False)
        self.subBanner.SetActive(False)
        self.start_Button.SetActive(False)
        self.start_Button.text.SetActive(False)
        self.rank_Graph_Button.SetActive(False)
        self.rank_Graph_Button.text.SetActive(False)
        self.exit_Button.SetActive(False)
        self.exit_Button.text.SetActive(False)
        pass

    def Handle_Event(self):
        if self.start_Button.isClick is True:
            self.SetActive(False)
            ModePopUp.Instance.SetActive(True)
            self.start_Button.isClick = False
            self.start_Button.frameCount = 0
        elif self.exit_Button.isClick is True:
            game_framework.quit()
        elif self.rank_Graph_Button.isClick is True:
            game_framework.push_state(rank_state)
            self.rank_Graph_Button.isClick = False
            self.rank_Graph_Button.frameCount = 0
        pass
    pass

class ModePopUp(Object):
    Instance = None
    def __init__(self):
        super(ModePopUp, self).__init__()
        self.frame = Object()
        self.frame.image = load_image('image/UI/Frame_White.png')
        self.frame.image_type = [0, 0, 600, 600]
        self.frame.transform.Position += Instance.windowSize / 2
        self.frame.transform.Scale *= 1.2

        self.modeButton = dict()
        self.modeButton['쉬움'] = Button(1)
        self.modeButton['쉬움'].name = 'Easy'
        self.modeButton['보통'] = Button(1)
        self.modeButton['보통'].name = 'Normal'
        self.modeButton['어려움'] = Button(1)
        self.modeButton['어려움'].name = 'Hard'

        i = 0
        file = load_workbook('Data/LevelingData.xlsx', data_only=True)
        for key, value in self.modeButton.items():
            sheet = file[value.name]
            value.transform.Scale *= 2
            value.transform.Position += [Instance.windowSize[0] // 2, 650 - i * 200]
            # Text
            text = Text(40)
            text.type = 'Middle'
            text.text = key
            text.transform.Position = value.transform.Position
            value.text = text
            # 설명 Text
            explainText = Text(30)
            explainText.font = explainText.fontList['Ghanachocolate']
            explainText.lineSpacing = 40
            explainText.text = []
            explainText.color = [255,255,0]

            for n in range(sheet["E7"].value):
                explainText.text.append(sheet.cell(row=n + 7, column=6).value)
            value.explainText = explainText

            i += 1
            pass
        self.textFrame = Object()
        self.textFrame.image = load_image('image/UI/Frame.png')
        self.textFrame.image_type = [0, 0, 600, 600]
        self.textFrame.transform.Scale *= [1.2, 0.25]
        self.textFrame.SetActive(False)

        ui_Render.AddObject(self.frame, 0)
        ui_Render.AddObject(self.textFrame, 3)
        ModePopUp.Instance = self
        pass

    def Enable(self):
        self.frame.SetActive(True)
        for mod in self.modeButton.values():
            mod.SetActive(True)
            mod.text.SetActive(True)
            mod.explainText.SetActive(True)
        pass

    def Disable(self):
        self.frame.SetActive(False)
        self.textFrame.SetActive(False)
        for mod in self.modeButton.values():
            mod.SetActive(False)
            mod.text.SetActive(False)
            mod.explainText.SetActive(False)
        pass

    def Update(self):
        self.textFrame.SetActive(False)
        for value in self.modeButton.values():
            if value.isOnMouse:
                value.explainText.transform.Position = value.mousePos + [35, -25]
                value.explainText.SetActive(True)
                self.textFrame.SetActive(True)
                self.textFrame.transform.Position = value.mousePos + [350, -65]
            else:
                value.explainText.SetActive(False)
            pass
        pass

    def Handle_Event(self):
        for value in self.modeButton.values():
            if value.isClick:
                game_state.level = value.name
                game_framework.change_state(game_state)
            pass
        for event in Object.events:
            if event.type == SDL_KEYDOWN:
                if event.key == SDLK_ESCAPE:
                    self.SetActive(False)
                    BannerPopUp.Instance.SetActive(True)
                    return
        pass
